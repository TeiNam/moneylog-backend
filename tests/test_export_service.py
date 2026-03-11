"""
ExportService 단위 테스트.

CSV 내보내기(UTF-8 BOM, 헤더, 데이터), 엑셀 내보내기(시트명, 헤더, 데이터),
카테고리/영역 필터링, 빈 데이터, 비밀 거래 포함을 검증한다.
Requirements: 9.1~9.9, 10.1~10.8
"""

import csv
import io
from datetime import date

import pytest
from openpyxl import load_workbook

from app.repositories.asset_repository import AssetRepository
from app.repositories.transaction_repository import TransactionRepository
from app.schemas.transaction import TransactionCreateData
from app.services.export_service import EXPORT_HEADERS, ExportService
from tests.conftest import create_test_user


# ──────────────────────────────────────────────
# 헬퍼 함수
# ──────────────────────────────────────────────


async def _create_transaction(db_session, user_id, **overrides):
    """테스트용 거래를 생성한다."""
    defaults = {
        "user_id": user_id,
        "date": date(2025, 6, 1),
        "area": "GENERAL",
        "type": "EXPENSE",
        "major_category": "식비",
        "minor_category": "외식",
        "description": "테스트 거래",
        "amount": 10000,
        "discount": 0,
        "actual_amount": 10000,
        "source": "MANUAL",
        "is_private": False,
    }
    defaults.update(overrides)
    repo = TransactionRepository(db_session)
    # Pydantic 모델로 변환하여 레포지토리 호출
    return await repo.create(TransactionCreateData(**defaults))


async def _create_asset(db_session, user_id, name="테스트자산", sort_order=1):
    """테스트용 자산을 생성한다."""
    repo = AssetRepository(db_session)
    return await repo.create({
        "user_id": user_id,
        "name": name,
        "asset_type": "BANK_ACCOUNT",
        "ownership": "PERSONAL",
        "balance": 1_000_000,
        "sort_order": sort_order,
    })


def _build_service(db_session):
    """ExportService 인스턴스를 생성한다."""
    return ExportService(
        transaction_repo=TransactionRepository(db_session),
        asset_repo=AssetRepository(db_session),
    )


def _parse_csv(csv_bytes: io.BytesIO) -> list[list[str]]:
    """CSV BytesIO를 파싱하여 행 리스트를 반환한다."""
    raw = csv_bytes.read()
    # UTF-8 BOM 제거
    if raw.startswith(b"\xef\xbb\xbf"):
        raw = raw[3:]
    text = raw.decode("utf-8")
    reader = csv.reader(io.StringIO(text))
    return list(reader)


def _parse_xlsx(xlsx_bytes: io.BytesIO):
    """xlsx BytesIO를 파싱하여 워크북을 반환한다."""
    return load_workbook(xlsx_bytes)


# ══════════════════════════════════════════════
# CSV 내보내기 테스트
# ══════════════════════════════════════════════


@pytest.mark.asyncio
async def test_csv_export_utf8_bom(db_session):
    """CSV 내보내기 시 UTF-8 BOM이 포함되는지 검증한다. (요구사항 9.2)"""
    user = await create_test_user(db_session)
    service = _build_service(db_session)

    result = await service.export_csv(
        user=user,
        start_date=date(2025, 1, 1),
        end_date=date(2025, 12, 31),
    )

    raw = result.read()
    assert raw.startswith(b"\xef\xbb\xbf"), "CSV 파일은 UTF-8 BOM으로 시작해야 한다"


@pytest.mark.asyncio
async def test_csv_export_headers(db_session):
    """CSV 내보내기 시 올바른 헤더가 포함되는지 검증한다. (요구사항 9.3)"""
    user = await create_test_user(db_session)
    service = _build_service(db_session)

    result = await service.export_csv(
        user=user,
        start_date=date(2025, 1, 1),
        end_date=date(2025, 12, 31),
    )

    rows = _parse_csv(result)
    assert len(rows) >= 1
    assert rows[0] == EXPORT_HEADERS


@pytest.mark.asyncio
async def test_csv_export_with_data(db_session):
    """CSV 내보내기 시 거래 데이터가 올바르게 포함되는지 검증한다. (요구사항 9.1)"""
    user = await create_test_user(db_session)
    asset = await _create_asset(db_session, user.id, name="신한카드")
    await _create_transaction(
        db_session,
        user.id,
        date=date(2025, 6, 15),
        area="GENERAL",
        type="EXPENSE",
        major_category="식비",
        minor_category="외식",
        description="점심식사",
        amount=15000,
        discount=1000,
        actual_amount=14000,
        asset_id=asset.id,
        memo="맛있었다",
    )
    service = _build_service(db_session)

    result = await service.export_csv(
        user=user,
        start_date=date(2025, 6, 1),
        end_date=date(2025, 6, 30),
    )

    rows = _parse_csv(result)
    # 헤더 + 데이터 1행
    assert len(rows) == 2
    data_row = rows[1]
    assert data_row[0] == "2025-06-15"  # 날짜
    assert data_row[1] == "GENERAL"  # 영역
    assert data_row[2] == "EXPENSE"  # 유형
    assert data_row[3] == "식비"  # 대분류
    assert data_row[4] == "외식"  # 소분류
    assert data_row[5] == "점심식사"  # 설명
    assert data_row[6] == "15000"  # 금액
    assert data_row[7] == "1000"  # 할인
    assert data_row[8] == "14000"  # 실지출
    assert data_row[9] == "신한카드"  # 결제수단
    assert data_row[10] == "맛있었다"  # 메모


# ══════════════════════════════════════════════
# 엑셀 내보내기 테스트
# ══════════════════════════════════════════════


@pytest.mark.asyncio
async def test_xlsx_export_sheet_name(db_session):
    """엑셀 내보내기 시 시트명이 '거래내역'인지 검증한다. (요구사항 10.2)"""
    user = await create_test_user(db_session)
    service = _build_service(db_session)

    result = await service.export_xlsx(
        user=user,
        start_date=date(2025, 1, 1),
        end_date=date(2025, 12, 31),
    )

    wb = _parse_xlsx(result)
    assert wb.active.title == "거래내역"


@pytest.mark.asyncio
async def test_xlsx_export_headers(db_session):
    """엑셀 내보내기 시 올바른 헤더가 포함되는지 검증한다. (요구사항 10.3)"""
    user = await create_test_user(db_session)
    service = _build_service(db_session)

    result = await service.export_xlsx(
        user=user,
        start_date=date(2025, 1, 1),
        end_date=date(2025, 12, 31),
    )

    wb = _parse_xlsx(result)
    ws = wb.active
    header_row = [cell.value for cell in ws[1]]
    assert header_row == EXPORT_HEADERS


@pytest.mark.asyncio
async def test_xlsx_export_with_data(db_session):
    """엑셀 내보내기 시 거래 데이터가 올바르게 포함되는지 검증한다. (요구사항 10.1)"""
    user = await create_test_user(db_session)
    asset = await _create_asset(db_session, user.id, name="국민카드")
    await _create_transaction(
        db_session,
        user.id,
        date=date(2025, 7, 10),
        area="CAR",
        type="EXPENSE",
        major_category="교통",
        minor_category="주유",
        description="주유소",
        amount=80000,
        discount=5000,
        actual_amount=75000,
        asset_id=asset.id,
        memo="경유",
    )
    service = _build_service(db_session)

    result = await service.export_xlsx(
        user=user,
        start_date=date(2025, 7, 1),
        end_date=date(2025, 7, 31),
    )

    wb = _parse_xlsx(result)
    ws = wb.active
    # 헤더(1행) + 데이터(1행) = 총 2행
    assert ws.max_row == 2
    data_row = [cell.value for cell in ws[2]]
    assert data_row[0] == "2025-07-10"  # 날짜
    assert data_row[1] == "CAR"  # 영역
    assert data_row[2] == "EXPENSE"  # 유형
    assert data_row[3] == "교통"  # 대분류
    assert data_row[4] == "주유"  # 소분류
    assert data_row[5] == "주유소"  # 설명
    assert data_row[6] == 80000  # 금액
    assert data_row[7] == 5000  # 할인
    assert data_row[8] == 75000  # 실지출
    assert data_row[9] == "국민카드"  # 결제수단
    assert data_row[10] == "경유"  # 메모


# ══════════════════════════════════════════════
# 카테고리/영역 필터링 테스트
# ══════════════════════════════════════════════


@pytest.mark.asyncio
async def test_csv_export_category_filter(db_session):
    """CSV 내보내기 시 카테고리 필터가 올바르게 적용되는지 검증한다. (요구사항 9.4)"""
    user = await create_test_user(db_session)
    # 식비 거래
    await _create_transaction(
        db_session, user.id,
        major_category="식비", minor_category="외식", description="점심",
    )
    # 교통 거래
    await _create_transaction(
        db_session, user.id,
        major_category="교통", minor_category="버스", description="출근",
    )
    service = _build_service(db_session)

    result = await service.export_csv(
        user=user,
        start_date=date(2025, 1, 1),
        end_date=date(2025, 12, 31),
        category="식비",
    )

    rows = _parse_csv(result)
    # 헤더 + 식비 거래 1건
    assert len(rows) == 2
    assert rows[1][3] == "식비"


@pytest.mark.asyncio
async def test_csv_export_area_filter(db_session):
    """CSV 내보내기 시 영역 필터가 올바르게 적용되는지 검증한다. (요구사항 9.5)"""
    user = await create_test_user(db_session)
    # GENERAL 영역 거래
    await _create_transaction(
        db_session, user.id,
        area="GENERAL", description="일반거래",
    )
    # CAR 영역 거래
    await _create_transaction(
        db_session, user.id,
        area="CAR", description="차량거래",
    )
    service = _build_service(db_session)

    result = await service.export_csv(
        user=user,
        start_date=date(2025, 1, 1),
        end_date=date(2025, 12, 31),
        area="CAR",
    )

    rows = _parse_csv(result)
    # 헤더 + CAR 거래 1건
    assert len(rows) == 2
    assert rows[1][1] == "CAR"


@pytest.mark.asyncio
async def test_xlsx_export_category_filter(db_session):
    """엑셀 내보내기 시 카테고리 필터가 올바르게 적용되는지 검증한다. (요구사항 10.4)"""
    user = await create_test_user(db_session)
    await _create_transaction(
        db_session, user.id,
        major_category="식비", description="점심",
    )
    await _create_transaction(
        db_session, user.id,
        major_category="교통", description="출근",
    )
    service = _build_service(db_session)

    result = await service.export_xlsx(
        user=user,
        start_date=date(2025, 1, 1),
        end_date=date(2025, 12, 31),
        category="교통",
    )

    wb = _parse_xlsx(result)
    ws = wb.active
    # 헤더(1행) + 교통 거래(1행)
    assert ws.max_row == 2
    assert ws[2][3].value == "교통"


@pytest.mark.asyncio
async def test_xlsx_export_area_filter(db_session):
    """엑셀 내보내기 시 영역 필터가 올바르게 적용되는지 검증한다. (요구사항 10.5)"""
    user = await create_test_user(db_session)
    await _create_transaction(
        db_session, user.id,
        area="GENERAL", description="일반",
    )
    await _create_transaction(
        db_session, user.id,
        area="SUBSCRIPTION", description="구독",
    )
    service = _build_service(db_session)

    result = await service.export_xlsx(
        user=user,
        start_date=date(2025, 1, 1),
        end_date=date(2025, 12, 31),
        area="SUBSCRIPTION",
    )

    wb = _parse_xlsx(result)
    ws = wb.active
    assert ws.max_row == 2
    assert ws[2][1].value == "SUBSCRIPTION"


# ══════════════════════════════════════════════
# 빈 데이터 테스트
# ══════════════════════════════════════════════


@pytest.mark.asyncio
async def test_csv_export_empty_data(db_session):
    """거래가 없을 때 CSV에 헤더만 포함되는지 검증한다. (요구사항 9.9)"""
    user = await create_test_user(db_session)
    service = _build_service(db_session)

    result = await service.export_csv(
        user=user,
        start_date=date(2025, 1, 1),
        end_date=date(2025, 12, 31),
    )

    rows = _parse_csv(result)
    assert len(rows) == 1  # 헤더만
    assert rows[0] == EXPORT_HEADERS


@pytest.mark.asyncio
async def test_xlsx_export_empty_data(db_session):
    """거래가 없을 때 엑셀에 헤더만 포함되는지 검증한다. (요구사항 10.8)"""
    user = await create_test_user(db_session)
    service = _build_service(db_session)

    result = await service.export_xlsx(
        user=user,
        start_date=date(2025, 1, 1),
        end_date=date(2025, 12, 31),
    )

    wb = _parse_xlsx(result)
    ws = wb.active
    assert ws.max_row == 1  # 헤더만
    header_row = [cell.value for cell in ws[1]]
    assert header_row == EXPORT_HEADERS


# ══════════════════════════════════════════════
# 비밀 거래 포함 테스트
# ══════════════════════════════════════════════


@pytest.mark.asyncio
async def test_csv_export_includes_private_transactions(db_session):
    """CSV 내보내기 시 본인의 비밀 거래가 포함되는지 검증한다. (요구사항 9.7)"""
    user = await create_test_user(db_session)
    # 공개 거래
    await _create_transaction(
        db_session, user.id,
        description="공개거래", is_private=False,
    )
    # 비밀 거래
    await _create_transaction(
        db_session, user.id,
        description="비밀거래", is_private=True,
    )
    service = _build_service(db_session)

    result = await service.export_csv(
        user=user,
        start_date=date(2025, 1, 1),
        end_date=date(2025, 12, 31),
    )

    rows = _parse_csv(result)
    # 헤더 + 공개 거래 + 비밀 거래 = 3행
    assert len(rows) == 3
    descriptions = {row[5] for row in rows[1:]}
    assert "공개거래" in descriptions
    assert "비밀거래" in descriptions


@pytest.mark.asyncio
async def test_xlsx_export_includes_private_transactions(db_session):
    """엑셀 내보내기 시 본인의 비밀 거래가 포함되는지 검증한다. (요구사항 10.6 관련)"""
    user = await create_test_user(db_session)
    await _create_transaction(
        db_session, user.id,
        description="공개거래", is_private=False,
    )
    await _create_transaction(
        db_session, user.id,
        description="비밀거래", is_private=True,
    )
    service = _build_service(db_session)

    result = await service.export_xlsx(
        user=user,
        start_date=date(2025, 1, 1),
        end_date=date(2025, 12, 31),
    )

    wb = _parse_xlsx(result)
    ws = wb.active
    # 헤더(1행) + 공개 거래 + 비밀 거래 = 3행
    assert ws.max_row == 3
    descriptions = {ws.cell(row=r, column=6).value for r in range(2, 4)}
    assert "공개거래" in descriptions
    assert "비밀거래" in descriptions


# ══════════════════════════════════════════════
# 사용자 격리 테스트
# ══════════════════════════════════════════════


@pytest.mark.asyncio
async def test_csv_export_only_own_data(db_session):
    """CSV 내보내기 시 본인 거래만 포함되는지 검증한다. (요구사항 9.6)"""
    user_a = await create_test_user(db_session, email="a@test.com", nickname="유저A")
    user_b = await create_test_user(db_session, email="b@test.com", nickname="유저B")
    await _create_transaction(db_session, user_a.id, description="A의거래")
    await _create_transaction(db_session, user_b.id, description="B의거래")
    service = _build_service(db_session)

    result = await service.export_csv(
        user=user_a,
        start_date=date(2025, 1, 1),
        end_date=date(2025, 12, 31),
    )

    rows = _parse_csv(result)
    # 헤더 + A의 거래 1건
    assert len(rows) == 2
    assert rows[1][5] == "A의거래"


@pytest.mark.asyncio
async def test_xlsx_export_only_own_data(db_session):
    """엑셀 내보내기 시 본인 거래만 포함되는지 검증한다. (요구사항 10.6)"""
    user_a = await create_test_user(db_session, email="a@test.com", nickname="유저A")
    user_b = await create_test_user(db_session, email="b@test.com", nickname="유저B")
    await _create_transaction(db_session, user_a.id, description="A의거래")
    await _create_transaction(db_session, user_b.id, description="B의거래")
    service = _build_service(db_session)

    result = await service.export_xlsx(
        user=user_a,
        start_date=date(2025, 1, 1),
        end_date=date(2025, 12, 31),
    )

    wb = _parse_xlsx(result)
    ws = wb.active
    assert ws.max_row == 2
    assert ws[2][5].value == "A의거래"
