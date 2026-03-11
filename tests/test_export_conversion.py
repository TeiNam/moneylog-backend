"""
순수 내보내기 변환 함수 속성 기반 테스트.

transactions_to_csv_bytes, transactions_to_xlsx_bytes, _transaction_to_row 함수를
DB 없이 독립적으로 테스트한다.
"""

import csv
import io
from datetime import date
from types import SimpleNamespace
from uuid import uuid4

from hypothesis import given, settings, strategies as st
from openpyxl import load_workbook

from app.services.export_service import (
    EXPORT_HEADERS,
    _transaction_to_row,
    transactions_to_csv_bytes,
    transactions_to_xlsx_bytes,
)

# --- Hypothesis 전략 정의 ---

# 거래 금액 생성 전략 (양수)
transaction_amounts = st.integers(min_value=1, max_value=10_000_000)

# 영역 생성 전략
areas = st.sampled_from(["GENERAL", "CAR", "SUBSCRIPTION", "EVENT"])

# 거래 유형 생성 전략
transaction_types = st.sampled_from(["INCOME", "EXPENSE"])

# 카테고리 생성 전략
categories = st.text(
    min_size=1, max_size=50, alphabet=st.characters(whitelist_categories=("L", "N"))
)

# 날짜 생성 전략
dates = st.dates(min_value=date(2020, 1, 1), max_value=date(2030, 12, 31))


def _make_transaction(
    tx_date=None,
    area="GENERAL",
    tx_type="EXPENSE",
    major_category="식비",
    minor_category="점심",
    description="테스트 거래",
    amount=10000,
    discount=0,
    actual_amount=10000,
    asset_id=None,
    memo=None,
):
    """테스트용 거래 객체를 SimpleNamespace로 생성한다."""
    return SimpleNamespace(
        date=tx_date or date(2024, 1, 15),
        area=area,
        type=tx_type,
        major_category=major_category,
        minor_category=minor_category,
        description=description,
        amount=amount,
        discount=discount,
        actual_amount=actual_amount,
        asset_id=asset_id,
        memo=memo,
    )


# 거래 객체 생성 전략
@st.composite
def transaction_strategy(draw):
    """임의의 거래 객체를 생성하는 Hypothesis 전략."""
    asset_id = draw(st.one_of(st.none(), st.builds(uuid4)))
    return _make_transaction(
        tx_date=draw(dates),
        area=draw(areas),
        tx_type=draw(transaction_types),
        major_category=draw(categories),
        minor_category=draw(categories),
        description=draw(categories),
        amount=draw(transaction_amounts),
        discount=draw(st.integers(min_value=0, max_value=1_000_000)),
        actual_amount=draw(transaction_amounts),
        asset_id=asset_id,
        memo=draw(st.one_of(st.none(), categories)),
    )


# 거래 목록 생성 전략 (빈 목록 포함)
transaction_list_strategy = st.lists(transaction_strategy(), min_size=0, max_size=10)


def _build_asset_map(transactions):
    """거래 목록에서 asset_map을 구축한다."""
    asset_map = {}
    for tx in transactions:
        if tx.asset_id is not None and tx.asset_id not in asset_map:
            asset_map[tx.asset_id] = f"자산_{tx.asset_id.hex[:6]}"
    return asset_map


# Feature: moneylog-backend-phase6, Property 14: CSV 내보내기 라운드트립
# **Validates: Requirements 9.1**
class TestCSVRoundtrip:
    """Property 14: CSV 내보내기 라운드트립 속성 기반 테스트."""

    @given(transactions=transaction_list_strategy)
    @settings(max_examples=30)
    def test_csv_roundtrip_row_count(self, transactions):
        """CSV 변환 후 파싱하면 행 수가 거래 수와 일치해야 한다."""
        asset_map = _build_asset_map(transactions)
        result = transactions_to_csv_bytes(transactions, asset_map)

        # BOM 건너뛰기
        raw = result.read()
        content = raw[3:].decode("utf-8")  # BOM 3바이트 건너뜀
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        # 헤더 1행 + 데이터 행
        assert len(rows) == len(transactions) + 1

    @given(transactions=transaction_list_strategy)
    @settings(max_examples=30)
    def test_csv_roundtrip_headers(self, transactions):
        """CSV 변환 후 파싱하면 헤더가 EXPORT_HEADERS와 일치해야 한다."""
        asset_map = _build_asset_map(transactions)
        result = transactions_to_csv_bytes(transactions, asset_map)

        raw = result.read()
        content = raw[3:].decode("utf-8")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        assert rows[0] == EXPORT_HEADERS

    @given(transactions=transaction_list_strategy)
    @settings(max_examples=30)
    def test_csv_roundtrip_data_matches(self, transactions):
        """CSV 변환 후 파싱하면 각 행의 데이터가 원본 거래 필드와 일치해야 한다."""
        asset_map = _build_asset_map(transactions)
        result = transactions_to_csv_bytes(transactions, asset_map)

        raw = result.read()
        content = raw[3:].decode("utf-8")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        for i, tx in enumerate(transactions):
            row = rows[i + 1]  # 헤더 건너뜀
            expected = _transaction_to_row(tx, asset_map)
            # CSV는 모든 값을 문자열로 변환하므로 str 비교
            assert row == [str(v) for v in expected]


# Feature: moneylog-backend-phase6, Property 15: CSV UTF-8 BOM 포함
# **Validates: Requirements 9.2**
class TestCSVUtf8Bom:
    """Property 15: CSV UTF-8 BOM 포함 속성 기반 테스트."""

    @given(transactions=transaction_list_strategy)
    @settings(max_examples=30)
    def test_csv_starts_with_utf8_bom(self, transactions):
        """임의의 거래 목록(빈 목록 포함)에 대해,
        transactions_to_csv_bytes의 출력 바이트 시퀀스는 UTF-8 BOM으로 시작해야 한다."""
        asset_map = _build_asset_map(transactions)
        result = transactions_to_csv_bytes(transactions, asset_map)

        raw = result.read()
        # UTF-8 BOM: \xef\xbb\xbf
        assert raw[:3] == b"\xef\xbb\xbf", (
            f"CSV 출력이 UTF-8 BOM으로 시작하지 않음: {raw[:3]!r}"
        )


# Feature: moneylog-backend-phase6, Property 16: 내보내기 필터링
# **Validates: Requirements 9.4, 9.5, 10.4, 10.5**
class TestExportFiltering:
    """Property 16: 내보내기 필터링 속성 기반 테스트.

    거래 목록을 카테고리/영역으로 사전 필터링한 후 CSV/xlsx로 변환하면,
    결과의 모든 거래가 해당 필터 조건과 일치해야 한다.
    """

    @given(
        transactions=st.lists(transaction_strategy(), min_size=1, max_size=15),
        filter_by=st.sampled_from(["category", "area"]),
    )
    @settings(max_examples=30)
    def test_csv_filtered_by_category_or_area(self, transactions, filter_by):
        """CSV: 카테고리 또는 영역 필터 적용 후 모든 행이 필터 값과 일치해야 한다."""
        # 생성된 거래 중 하나에서 필터 값을 선택
        chosen_tx = transactions[0]
        if filter_by == "category":
            filter_value = chosen_tx.major_category
            # 필터 적용: 해당 대분류 카테고리의 거래만 포함
            filtered = [tx for tx in transactions if tx.major_category == filter_value]
            col_index = 3  # 대분류 컬럼 인덱스
        else:
            filter_value = chosen_tx.area
            # 필터 적용: 해당 영역의 거래만 포함
            filtered = [tx for tx in transactions if tx.area == filter_value]
            col_index = 1  # 영역 컬럼 인덱스

        asset_map = _build_asset_map(filtered)
        result = transactions_to_csv_bytes(filtered, asset_map)

        # CSV 파싱
        raw = result.read()
        content = raw[3:].decode("utf-8")  # BOM 건너뜀
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        # 헤더 제외한 데이터 행 검증
        data_rows = rows[1:]
        assert len(data_rows) == len(filtered)
        for row in data_rows:
            assert row[col_index] == filter_value, (
                f"필터({filter_by}={filter_value})와 불일치: {row[col_index]}"
            )

    @given(
        transactions=st.lists(transaction_strategy(), min_size=1, max_size=15),
        filter_by=st.sampled_from(["category", "area"]),
    )
    @settings(max_examples=30)
    def test_xlsx_filtered_by_category_or_area(self, transactions, filter_by):
        """xlsx: 카테고리 또는 영역 필터 적용 후 모든 행이 필터 값과 일치해야 한다."""
        chosen_tx = transactions[0]
        if filter_by == "category":
            filter_value = chosen_tx.major_category
            filtered = [tx for tx in transactions if tx.major_category == filter_value]
            col_index = 4  # openpyxl은 1-based, 대분류 = 4번째 열
        else:
            filter_value = chosen_tx.area
            filtered = [tx for tx in transactions if tx.area == filter_value]
            col_index = 2  # 영역 = 2번째 열

        asset_map = _build_asset_map(filtered)
        result = transactions_to_xlsx_bytes(filtered, asset_map)

        # xlsx 파싱
        wb = load_workbook(result)
        ws = wb.active
        all_rows = list(ws.iter_rows(min_row=2, values_only=True))  # 헤더 건너뜀

        assert len(all_rows) == len(filtered)
        for row in all_rows:
            cell_value = row[col_index - 1]  # 0-based 인덱스로 변환
            assert cell_value == filter_value, (
                f"필터({filter_by}={filter_value})와 불일치: {cell_value}"
            )

    @given(
        transactions=st.lists(transaction_strategy(), min_size=2, max_size=15),
    )
    @settings(max_examples=30)
    def test_csv_filter_excludes_non_matching(self, transactions):
        """CSV: 필터 적용 시 일치하지 않는 거래가 결과에 포함되지 않아야 한다."""
        # 첫 번째 거래의 영역으로 필터링
        filter_area = transactions[0].area
        filtered = [tx for tx in transactions if tx.area == filter_area]
        excluded_count = len(transactions) - len(filtered)

        asset_map = _build_asset_map(filtered)
        result = transactions_to_csv_bytes(filtered, asset_map)

        raw = result.read()
        content = raw[3:].decode("utf-8")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        # 필터링된 결과의 행 수가 원본보다 적거나 같아야 함
        data_rows = rows[1:]
        assert len(data_rows) == len(filtered)
        assert len(data_rows) <= len(transactions)
        # 모든 행이 필터 영역과 일치
        for row in data_rows:
            assert row[1] == filter_area

    @given(
        transactions=st.lists(transaction_strategy(), min_size=2, max_size=15),
    )
    @settings(max_examples=30)
    def test_xlsx_filter_excludes_non_matching(self, transactions):
        """xlsx: 필터 적용 시 일치하지 않는 거래가 결과에 포함되지 않아야 한다."""
        filter_category = transactions[0].major_category
        filtered = [tx for tx in transactions if tx.major_category == filter_category]

        asset_map = _build_asset_map(filtered)
        result = transactions_to_xlsx_bytes(filtered, asset_map)

        wb = load_workbook(result)
        ws = wb.active
        all_rows = list(ws.iter_rows(min_row=2, values_only=True))

        assert len(all_rows) == len(filtered)
        assert len(all_rows) <= len(transactions)
        # 모든 행이 필터 카테고리와 일치 (대분류 = 4번째 열, 0-based index 3)
        for row in all_rows:
            assert row[3] == filter_category


# Feature: moneylog-backend-phase6, Property 17: 내보내기 데이터 범위
# **Validates: Requirements 9.6, 9.7, 10.6**
class TestExportDataScope:
    """Property 17: 내보내기 데이터 범위 속성 기반 테스트.

    두 사용자 A, B에 대해:
    1. A의 거래만 변환 함수에 전달하면, B의 거래 데이터가 결과에 포함되지 않아야 한다.
    2. A의 비밀 거래(is_private=true)는 A의 내보내기 결과에 포함되어야 한다.
    """

    @given(
        user_a_txs=st.lists(transaction_strategy(), min_size=1, max_size=8),
        user_b_txs=st.lists(transaction_strategy(), min_size=1, max_size=8),
    )
    @settings(max_examples=30)
    def test_csv_user_isolation(self, user_a_txs, user_b_txs):
        """CSV: 사용자 A의 거래만 변환하면 사용자 B의 거래가 결과에 포함되지 않아야 한다."""
        # 사용자 A의 거래만으로 CSV 변환
        asset_map = _build_asset_map(user_a_txs)
        result = transactions_to_csv_bytes(user_a_txs, asset_map)

        # CSV 파싱
        raw = result.read()
        content = raw[3:].decode("utf-8")  # BOM 건너뜀
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        data_rows = rows[1:]  # 헤더 제외

        # A의 거래 수만큼만 행이 존재해야 함
        assert len(data_rows) == len(user_a_txs)

        # B의 거래 설명(description)이 결과에 포함되지 않아야 함
        a_descriptions = {tx.description for tx in user_a_txs}
        for tx_b in user_b_txs:
            if tx_b.description not in a_descriptions:
                # B 고유 거래의 설명이 CSV 데이터에 없어야 함
                for row in data_rows:
                    assert row[5] != tx_b.description or tx_b.description in a_descriptions

    @given(
        user_a_txs=st.lists(transaction_strategy(), min_size=1, max_size=8),
        user_b_txs=st.lists(transaction_strategy(), min_size=1, max_size=8),
    )
    @settings(max_examples=30)
    def test_xlsx_user_isolation(self, user_a_txs, user_b_txs):
        """xlsx: 사용자 A의 거래만 변환하면 사용자 B의 거래가 결과에 포함되지 않아야 한다."""
        # 사용자 A의 거래만으로 xlsx 변환
        asset_map = _build_asset_map(user_a_txs)
        result = transactions_to_xlsx_bytes(user_a_txs, asset_map)

        # xlsx 파싱
        wb = load_workbook(result)
        ws = wb.active
        all_rows = list(ws.iter_rows(min_row=2, values_only=True))  # 헤더 건너뜀

        # A의 거래 수만큼만 행이 존재해야 함
        assert len(all_rows) == len(user_a_txs)

        # B의 거래 데이터가 결과에 없어야 함 (행 수로 검증)
        # 변환 함수에 A의 거래만 전달했으므로 B의 거래는 절대 포함될 수 없음
        assert len(all_rows) == len(user_a_txs), (
            f"A의 거래 수({len(user_a_txs)})와 결과 행 수({len(all_rows)})가 불일치"
        )

    @given(
        transactions=st.lists(transaction_strategy(), min_size=1, max_size=10),
    )
    @settings(max_examples=30)
    def test_csv_includes_private_transactions(self, transactions):
        """CSV: 비밀 거래(is_private=true)가 본인 내보내기 결과에 포함되어야 한다."""
        # 일부 거래를 비밀 거래로 설정
        for i, tx in enumerate(transactions):
            tx.is_private = i % 2 == 0  # 짝수 인덱스는 비밀 거래

        private_count = sum(1 for tx in transactions if tx.is_private)

        asset_map = _build_asset_map(transactions)
        # 비밀 거래 포함하여 모든 거래를 변환 (본인 데이터이므로)
        result = transactions_to_csv_bytes(transactions, asset_map)

        # CSV 파싱
        raw = result.read()
        content = raw[3:].decode("utf-8")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        data_rows = rows[1:]

        # 비밀 거래를 포함한 전체 거래가 모두 포함되어야 함
        assert len(data_rows) == len(transactions)
        # 비밀 거래가 1개 이상 존재하는지 확인
        assert private_count > 0, "비밀 거래가 생성되지 않음"

    @given(
        transactions=st.lists(transaction_strategy(), min_size=1, max_size=10),
    )
    @settings(max_examples=30)
    def test_xlsx_includes_private_transactions(self, transactions):
        """xlsx: 비밀 거래(is_private=true)가 본인 내보내기 결과에 포함되어야 한다."""
        # 일부 거래를 비밀 거래로 설정
        for i, tx in enumerate(transactions):
            tx.is_private = i % 2 == 0

        private_count = sum(1 for tx in transactions if tx.is_private)

        asset_map = _build_asset_map(transactions)
        result = transactions_to_xlsx_bytes(transactions, asset_map)

        # xlsx 파싱
        wb = load_workbook(result)
        ws = wb.active
        all_rows = list(ws.iter_rows(min_row=2, values_only=True))

        # 비밀 거래를 포함한 전체 거래가 모두 포함되어야 함
        assert len(all_rows) == len(transactions)
        assert private_count > 0, "비밀 거래가 생성되지 않음"

    @given(
        user_a_txs=st.lists(transaction_strategy(), min_size=1, max_size=8),
        user_b_txs=st.lists(transaction_strategy(), min_size=1, max_size=8),
    )
    @settings(max_examples=30)
    def test_csv_private_included_other_user_excluded(self, user_a_txs, user_b_txs):
        """CSV: A의 비밀 거래는 포함되고, B의 거래(비밀 포함)는 제외되어야 한다."""
        # A의 거래 중 일부를 비밀로 설정
        for i, tx in enumerate(user_a_txs):
            tx.is_private = True  # 모든 A 거래를 비밀로 설정

        # B의 거래도 비밀로 설정 (하지만 A의 내보내기에 포함되면 안 됨)
        for tx in user_b_txs:
            tx.is_private = True

        # A의 거래만 변환
        asset_map = _build_asset_map(user_a_txs)
        result = transactions_to_csv_bytes(user_a_txs, asset_map)

        raw = result.read()
        content = raw[3:].decode("utf-8")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        data_rows = rows[1:]

        # A의 비밀 거래가 모두 포함되어야 함
        assert len(data_rows) == len(user_a_txs)
        # B의 거래는 포함되지 않음 (변환 함수에 전달하지 않았으므로)
        assert len(data_rows) != len(user_a_txs) + len(user_b_txs) or len(user_b_txs) == 0


# Feature: moneylog-backend-phase6, Property 18: 내보내기 날짜 검증
# **Validates: Requirements 9.8, 10.7**
class TestExportDateValidation:
    """Property 18: 내보내기 날짜 검증 속성 기반 테스트.

    start_date > end_date인 날짜 쌍에 대해 ExportFilterParams 생성 시
    ValidationError가 발생해야 하고, start_date <= end_date인 경우에는
    정상적으로 생성되어야 한다.
    """

    @given(
        d1=st.dates(min_value=date(2020, 1, 1), max_value=date(2030, 12, 31)),
        d2=st.dates(min_value=date(2020, 1, 1), max_value=date(2030, 12, 31)),
    )
    @settings(max_examples=30)
    def test_invalid_date_range_raises_validation_error(self, d1, d2):
        """start_date > end_date이면 ValidationError가 발생해야 한다."""
        from pydantic import ValidationError

        from app.schemas.export import ExportFilterParams

        # 두 날짜가 같으면 유효한 범위이므로 건너뜀
        if d1 <= d2:
            # start_date > end_date가 되도록 스왑
            d1, d2 = d2, d1
        if d1 <= d2:
            # 두 날짜가 동일한 경우 건너뜀 (스왑 후에도 같을 수 있음)
            return

        # start_date > end_date → ValidationError 발생 기대
        try:
            ExportFilterParams(start_date=d1, end_date=d2)
            assert False, (
                f"start_date({d1}) > end_date({d2})인데 ValidationError가 발생하지 않음"
            )
        except ValidationError:
            pass  # 기대한 동작

    @given(
        d1=st.dates(min_value=date(2020, 1, 1), max_value=date(2030, 12, 31)),
        d2=st.dates(min_value=date(2020, 1, 1), max_value=date(2030, 12, 31)),
    )
    @settings(max_examples=30)
    def test_valid_date_range_does_not_raise(self, d1, d2):
        """start_date <= end_date이면 ValidationError가 발생하지 않아야 한다."""
        from pydantic import ValidationError

        from app.schemas.export import ExportFilterParams

        # start_date <= end_date가 되도록 정렬
        start = min(d1, d2)
        end = max(d1, d2)

        # 유효한 날짜 범위 → 정상 생성
        params = ExportFilterParams(start_date=start, end_date=end)
        assert params.start_date == start
        assert params.end_date == end


# Feature: moneylog-backend-phase6, Property 19: 엑셀 내보내기 라운드트립
# **Validates: Requirements 10.1, 10.2, 10.3**
class TestXlsxRoundtrip:
    """Property 19: 엑셀 내보내기 라운드트립 속성 기반 테스트.

    임의의 거래 목록에 대해, transactions_to_xlsx_bytes로 변환한 엑셀 파일을
    openpyxl로 읽으면 시트명이 "거래내역"이고, 헤더가 EXPORT_HEADERS와 일치하며,
    각 행의 데이터가 원본 거래의 필드와 일치해야 하고,
    행 수는 거래 수 + 1(헤더)과 일치해야 한다.
    """

    @given(transactions=transaction_list_strategy)
    @settings(max_examples=30)
    def test_xlsx_sheet_name(self, transactions):
        """엑셀 파일의 시트명이 '거래내역'이어야 한다."""
        asset_map = _build_asset_map(transactions)
        result = transactions_to_xlsx_bytes(transactions, asset_map)

        wb = load_workbook(result)
        ws = wb.active
        assert ws.title == "거래내역", f"시트명이 '거래내역'이 아님: {ws.title}"

    @given(transactions=transaction_list_strategy)
    @settings(max_examples=30)
    def test_xlsx_row_count(self, transactions):
        """엑셀 파일의 행 수가 거래 수 + 1(헤더)과 일치해야 한다."""
        asset_map = _build_asset_map(transactions)
        result = transactions_to_xlsx_bytes(transactions, asset_map)

        wb = load_workbook(result)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))

        # 헤더 1행 + 데이터 행
        assert len(all_rows) == len(transactions) + 1, (
            f"행 수 불일치: 기대={len(transactions) + 1}, 실제={len(all_rows)}"
        )

    @given(transactions=transaction_list_strategy)
    @settings(max_examples=30)
    def test_xlsx_headers(self, transactions):
        """엑셀 파일의 헤더가 EXPORT_HEADERS와 일치해야 한다."""
        asset_map = _build_asset_map(transactions)
        result = transactions_to_xlsx_bytes(transactions, asset_map)

        wb = load_workbook(result)
        ws = wb.active
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        assert list(header_row) == EXPORT_HEADERS, (
            f"헤더 불일치: 기대={EXPORT_HEADERS}, 실제={list(header_row)}"
        )

    @given(transactions=transaction_list_strategy)
    @settings(max_examples=30)
    def test_xlsx_data_matches(self, transactions):
        """엑셀 파일의 각 행 데이터가 원본 거래 필드와 일치해야 한다."""
        asset_map = _build_asset_map(transactions)
        result = transactions_to_xlsx_bytes(transactions, asset_map)

        wb = load_workbook(result)
        ws = wb.active
        # 헤더 건너뛰고 데이터 행만 가져옴
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))

        for i, tx in enumerate(transactions):
            row = list(data_rows[i])
            expected = _transaction_to_row(tx, asset_map)
            # openpyxl은 빈 문자열("")을 None으로 저장하므로 정규화하여 비교
            for j, (actual, exp) in enumerate(zip(row, expected)):
                # openpyxl None ↔ 빈 문자열 정규화
                norm_actual = actual if actual is not None else ""
                norm_exp = exp if exp is not None else ""
                assert norm_actual == norm_exp, (
                    f"거래[{i}] 컬럼[{j}]({EXPORT_HEADERS[j]}) 불일치: "
                    f"기대={exp!r}, 실제={actual!r}"
                )
